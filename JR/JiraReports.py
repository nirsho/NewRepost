import matplotlib as mpl
import requests
import argparse
import configparser
import base64
import logging
import os
import sys
from logging.handlers import RotatingFileHandler
from requests.auth import HTTPBasicAuth
from datetime import datetime
from datetime import date
from datetime import timedelta
import calendar
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import io
from openpyxl.chart import LineChart, Reference
import openpyxl.utils.cell as utils
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta





"""
    JiraReports - basic html tool to fetch reports from Jira cloud based on jira RESTful API.

    usage: JiraReports [-h] [-s STARTDATE] [-e ENDDATE] [-m] [-p] [-r REPORTS [REPORTS ...]] [-b BREAKDOWN] [-g]

           optional arguments:
             -h, --help            show this help message and exit
             -s STARTDATE          Start date for the report in the format of : YYYY-MM-DD
             -e ENDDATE            End date for teh report in the format of : YYYY-MM-DD
             -m                    Print reports menu
             -p                    Flag for periodic reports (used for CRON jobs)
             -r REPORTS [REPORTS ...] List of reports seperated by space
             -b BREAKDOWN          Break Down resolution - 'd' for day 'w' for week and 'm' for month
             -g                    Flag to generate graphical reports
"""


class Set:
    """
        Set - helper class, contains one set of a series.
    """

    def __init__(self, name, color):
        self.name = name
        self.color = color
        self.x = list()
        self.y = list()
        self.avg = 0
        self.cnt = 0
        self.type = None

    def calcMetrics(self):
        """
            calcMetrics() - a function to calculate the count and the average of all values to the set y values
        """
        for elem in self.y:
            self.cnt = self.cnt + elem
        self.avg = self.cnt / len(self.y)


class Report:
    """
        Report - helper class contains all the series in a report
    """

    def __init__(self, name):
        self.name = name
        self.series = list()

    def addSet(self, set):
        """
        addSet - addes a set to the serries
        """
        set.calcMetrics()
        self.series.append(set)


def initLogger(logFile):
    logging.basicConfig(
        handlers=[RotatingFileHandler(logFile, maxBytes=1000000, backupCount=5)],
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s [%(name)s.%(funcName)s:%(lineno)d] %(message)s",
        datefmt='%Y-%m-%dT%H:%M:%S')
    log = logging.getLogger("JIRAREPORTS")
    log.setLevel(logging.DEBUG)

    return log


def printMenu():
    """
        Print the menu , used for cli operation
    """
    print("{0:<10}{1:<11}{2:<11}".format("Report #", "|", "Description"))
    print('-' * 32)
    print("{0:<10}{1:<11}{2:<11}".format("15", "|", "Create vs Resolved"))
    print("{0:<10}{1:<11}{2:<11}".format("17", "|", "SLA Met vs Breached"))
    print("{0:<10}{1:<11}{2:<11}".format("25", "|", "SLA - Time"))
    print("{0:<10}{1:<11}{2:<11}".format("26", "|", "Cr by Type"))
    print("{0:<10}{1:<11}{2:<11}".format("27", "|", "Work Load"))


def parseInput(startTime, endTime, breakDown, reports):
    if not startTime:
        startTime = input("Please insert start date for the report (please set input as : YYYY-MM-DD)\n")
    if not endTime:
        endTime = input("Please insert end date for teh report (please set input as : YYYY-MM-DD)\n")
    if not breakDown:
        breakDown = input("Please insert report break down (d - date, w - week, m - month)\n")
    if not reports:
        print("please insert report number , multiple report can be selected by seperating them with a comma \",\":")
        printMenu()
        reports = input("Report(s): ").split(',')
    return startTime, endTime, breakDown, reports


def checkDate(value):
    y, m, d = value.split('-')
    if ((date.today() - datetime.strptime(value, "%Y-%m-%d").date()).days < 0):
        raise argparse.ArgumentTypeError("%s date can only be smaller than today" % value)
    return value


# ARG parser
def parseArgs():
    parser = argparse.ArgumentParser(prog='JiraReports', description='Fetch reports from Jira.')
    parser.add_argument('-s', dest="startDate", type=checkDate,
                        help='Start date for the report in the format of : YYYY-MM-DD')
    parser.add_argument('-e', dest='endDate', type=str, action='store',
                        help='End date for teh report in the format of : YYYY-MM-DD')
    parser.add_argument('-m', dest='menu', action='store_true', help='Print reports menu')
    parser.add_argument('-p', dest='periodic', action='store_true',
                        help='Flag for periodic reports (used for CRON jobs)')
    parser.add_argument('-r', dest='reports', type=str, action='store', nargs='+',
                        help='List of reports seperated by space')
    parser.add_argument('-b', dest='breakDown', type=str, action='store',
                        help='Break Down resolution - \'d\' for day \'w\' for week and \'m\' for month')
    parser.add_argument('-g', dest='graph', action='store_true', default=False,
                        help='Flag to generate graphical reports')

    return parser.parse_args()


def retriveReports(uri, auth, headers, startTime, endTime, breakDown, reports):
    print(startTime + " " + endTime)
    log = logging.getLogger('JIRAREPORTS')
    log.info("Starting retriveReports from JIRA")
    result = list()
    for rep in reports:
        url = uri + rep + "/date-range?startDate=" + startTime + "&endDate=" + endTime + "&timeBreakdown=" + breakDown
        # print(url)
        response = requests.request("GET", url, headers=headers, auth=auth)
        if response.status_code != 200:
            log.info('some issu occured with the JIRA API')
            log.info(response.text)
            sys.exit(0)
        jsn = response.json()
        # print(jsn)
        # result.append(jsn)
        reportLocal = Report(jsn["name"])
        for ser in jsn['series']:
            seriesLocal = Set(ser['label'], ser['color'])
            seriesLocal.type = ser["seriesType"]["yaxis"]['typeKey'].split('.')[-1]
            for ln in ser['data']:
                # adjust the winter time time zone
                addTime = timedelta(hours=2)
                seriesLocal.x.append((datetime.fromtimestamp(ln['x'] / 1e3) + addTime).date())
                multVal = 1
                if (seriesLocal.type == "duration"):
                    multVal = (0.001 / 3600)
                seriesLocal.y.append(ln['y'] * multVal)
            reportLocal.addSet(seriesLocal)
        result.append(reportLocal)
    return result


def excelout(results, filename):
    # Create a workbook and remove the default sheet
    workbook = Workbook()
    workbook.remove(workbook.active)
    for result in results:
        sheet_name = str(result.name).replace("/", " or ")
        worksheet = workbook.create_sheet(title=sheet_name)
        data = {}
        reports = result.series
        values = []
        desired_width = 20  # Adjust this value as needed
        for report in reports:
            values.append(report.name)
            data[result.name] = values
            for i in range(len(report.x)):
                x = str(report.x[i])
                y = round(report.y[i], 2) if isinstance(report.y[i], float) else report.y[i]
                if x in data:
                    data[x].append(y)
                else:
                    data[x] = [y]
        df = pd.DataFrame(data)
        df = df.transpose().reset_index()
        df.columns = ['Date'] + list(df.columns[1:])

        # Write the DataFrame to the Excel sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
            worksheet.append(row)

        for col in worksheet.columns:
            worksheet.column_dimensions[utils.get_column_letter(col[0].column)].width = desired_width


        # Save workbook after all sheets are added
    workbook.save(filename)

    # Re-open the workbook to add charts
    workbook = load_workbook(filename)

    for result in results:
        sheet_name = str(result.name).replace("/", " or ")
        worksheet = workbook[sheet_name]
        chart = LineChart()
        chart.title = sheet_name

        for col in range(2, worksheet.max_column + 1):
            data = Reference(worksheet, min_col=col, min_row=1, max_row=worksheet.max_row, max_col=col)
            chart.add_data(data, titles_from_data=True)

        categories = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
        chart.set_categories(categories)
        worksheet.add_chart(chart, f"G{worksheet.max_row + 2}")  # Dynamically place chart
        chart.height = 8
        chart.width = 20

    # Save workbook after all sheets are added
    workbook.save(filename)


def plotGraph(results):
    mpl.style.use('seaborn')
    fig, axs = plt.subplots(len(results))
    if (len(results) == 1):
        axs = [axs]
    i = 0
    for rep in results:
        for ser in rep.series:
            if ser.type == 'duration':
                val = ser.avg
            else:
                val = ser.cnt
            axs[i].plot(ser.x, ser.y, 'o-', color=ser.color, label=ser.name + ' - ' + str(val))
        axs[i].legend(bbox_to_anchor=(1.01, 0.66), loc="upper left")
        axs[i].set_title(rep.name)
        i = i + 1
    plt.show()


def generateGraphs(results):
    graphs = dict()
    mpl.style.use('seaborn')
    for rep in results:
        plt.clf()
        for ser in rep.series:
            if ser.type == 'duration':
                val = ser.avg
            else:
                val = ser.cnt
            plt.plot(ser.x, ser.y, 'o-', color=ser.color, label=ser.name + ' - ' + str(val))
        buff = io.BytesIO()
        locs, labels = plt.xticks()
        plt.xticks(mpl.dates.date2num(ser.x), labels=ser.x, rotation=45)
        plt.savefig(buff, format='png', dpi=300, edgecolor='black', bbox_inches='tight')
        imgEncoded = base64.b64encode(buff.getvalue()).decode("utf-8").replace("\n", "")
        graphs[rep.name] = imgEncoded
    return graphs


def jsons_to_pdf(json_list, pdf_filename, start_date_str, end_date_str):
    plots_generated = False
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

    with PdfPages(pdf_filename) as pdf:
        # Welcome page setup
        plt.figure(figsize=(8, 11))  # Standard US Letter page size in inches
        plt.axis('off')  # Turn off the axis
        welcome_text = f"Report from {start_date_str} to {end_date_str}"
        plt.text(0.5, 0.5, welcome_text, fontsize=16, ha='center', va='center', transform=plt.gcf().transFigure)
        pdf.savefig()  # Save the welcome page
        plt.close()

        for json_obj in json_list:
            fig, ax = plt.subplots(figsize=(16, 8))  # Half page for the graph
            report_name = json_obj['name']
            time = "'yAxisType': 'time'" in str(json_obj['series'])
            y_axis_label = 'Minutes' if time else 'CRs'

            ax.set_title(report_name, fontsize=12)
            ax.set_xlabel('Date', fontsize=12)
            ax.set_ylabel(y_axis_label, fontsize=12)

            # Additional ax2 for the second Y axis if necessary
            ax2 = ax.twinx()

            # Collect total texts
            total_texts = []

            for series in json_obj['series']:
                label = series['label']
                color = series['color']
                linestyle = series.get('linestyle', '-')
                marker = series.get('marker', 'o')
                data = pd.DataFrame(series['data'])
                data['x'] = pd.to_datetime(data['x'], unit='ms')
                if time:
                    data['y'] = data['y'] / 60000  # Convert to minutes if necessary
                ax.plot(data['x'], data['y'], label=label, color=color, linestyle=linestyle, marker=marker)

                # Compute and format total for this series
                total = int(data['y'].sum())
                if time:
                    if total > 60:
                        minutes = total % 60
                        total = int(total / 60)
                        total_texts.append(f"Total for {label}: {total} Hours and " + str(minutes) + " Minutes")
                    else:
                        total_texts.append(f"Total for {label}: {total} Minutes")
                else:
                    total_texts.append(f"Total for {label}: {total} CRs")



            # Format and set the X and Y axis
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator())
            ax.set_xlim(start_date, end_date)
            ax.grid(True)
            ax.legend(loc='upper left')

            # Customize ax2 if used for dual Y-axis
            # ax2.set_ylabel('Second metric', fontsize=12)
            # ax2.legend(loc='upper right')

            fig.autofmt_xdate()

            # Save the graph page
            pdf.savefig(fig)
            plt.close(fig)

            # Text page for totals
            fig, ax = plt.subplots(figsize=(8, 5.5))  # Half page for the text
            plt.axis('off')  # Hide the axis for the text page
            text_y_position = 0.9
            for total_text in total_texts:
                ax.text(0.5, text_y_position, total_text, fontsize=12, ha='center', transform=ax.transAxes)
                text_y_position -= 0.1  # Move text down for each line

            # Save the text page
            pdf.savefig(fig)
            plt.close(fig)
            plots_generated = True

    if not plots_generated:
        print("No plots were generated. Please check the input data.")
    else:
        print(f"PDF file '{pdf_filename}' has been generated with the plots and data texts.")




def addMonthlyYTD(result, uri, auth, headers, startTime, endTime, reports):
    log = logging.getLogger('JIRAREPORTS')
    log.info("Starting YTD for Created VS resolved")
    rep = '15'
    reports.append(rep)
    # if int(endTime.split("-")[1]) >= 4:
    breakDown = 'm'
    # else:
    #     breakDown = 'w'
    print(reports)
    print(endTime)
    print(breakDown)
    if rep in reports:
        # url = uri + rep + "/date-range?startDate=" + startTime + "&endDate=" + endTime + "&timeBreakdown=" + breakDown

        url = uri + rep + "/date-range?startDate=" + startTime + "&endDate=" + endTime + "&timeBreakdown=" + breakDown
        print(url)
        response = requests.request("GET", url, headers=headers, auth=auth)
        if response.status_code != 200:
            log.info('some issu occured with the JIRA API')
            log.info(response.text)
            sys.exit(0)
        jsn = response.json()
        reportLocal = Report(jsn["name"])
        reportLocal.name = jsn["name"] + " by month"
        for ser in jsn['series']:
            seriesLocal = Set(ser['label'], ser['color'])
            seriesLocal.type = ser["seriesType"]["yaxis"]['typeKey'].split('.')[-1]
            for ln in ser['data']:
                # adjust the winter time time zone
                addTime = timedelta(hours=2)
                seriesLocal.x.append((datetime.fromtimestamp(ln['x'] / 1e3) + addTime).date())
                multVal = 1
                if (seriesLocal.type == "duration"):
                    multVal = (0.001 / 3600)
                seriesLocal.y.append(ln['y'] * multVal)
            reportLocal.addSet(seriesLocal)
        result.append(reportLocal)
    for x in result:
        print(x)
    return result


def main():
    dir = os.path.dirname(__file__)
    logFile = os.path.join(dir, 'log', 'JiraReports.log')
    log = initLogger(logFile)
    log.info("JiraReports is starting ...")
    args = parseArgs()
    if (args.menu):
        printMenu()
        return 0

    log.info('Starting config file parsing')
    cnf = configparser.ConfigParser()
    cnf.read('.\conf.ini')
    log.info('Finished config file parsing')
    auth = HTTPBasicAuth(cnf['JIRA']['user'], cnf['JIRA']['token'])
    headers = {"accept": "application/json"}
    url = cnf['JIRA']['url']
    if (args.periodic):
        log.info('Start periodic procedure')
        args.graph = False
        today = datetime.now().date()
        endYear = today.year
        endMonth = today.month
        if (endMonth == 1):
            startYear = endYear - 1
            startMonth = 12
            endYear = endYear -1
        else:
            startYear = endYear
            startMonth = endMonth - 1
        daysInMonth = calendar.monthrange(startYear, startMonth)[1]
        startTime = datetime(startYear, startMonth, 1).strftime("%Y-%m-%d")
        startTime_date = datetime(startYear, startMonth, 1)
        # Calculate six months ago
        six_months_ago = startTime_date - relativedelta(months=6)
        # Get the first day of the month six months ago
        first_day_six_months_ago = six_months_ago.replace(day=1)
        print(first_day_six_months_ago)
        first_day_six_months_ago = first_day_six_months_ago.strftime("%Y-%m-%d")
        endTime = datetime(endYear, startMonth, daysInMonth).strftime("%Y-%m-%d")
        print(startTime)
        print(endTime)
        breakDown = 'w'
        reports = cnf['JIRA']['defaultReports'].split(',')
        print(reports)
        results = retriveReports(url, auth, headers, startTime, endTime, breakDown, reports)
        results = addMonthlyYTD(results, url, auth, headers, first_day_six_months_ago, endTime, reports)



    else:
        startTime, endTime, breakDown, reports = parseInput(args.startDate, args.endDate, args.breakDown, args.reports)
        results = retriveReports(url, auth, headers, startTime, endTime, breakDown, reports)
        results = addMonthlyYTD(results, url, auth, headers, startTime, endTime, reports)



    excelout(results, "Jira Report " + startTime + " to " + endTime + ".xlsx")



if __name__ == "__main__":
    main()
    # try:
        # main()
    # except Exception as e:
    #     log = logging.getLogger('JIRAREPORTS')
    #     log.error("Unhandled exception occured:")
    #     log.exception(e)

